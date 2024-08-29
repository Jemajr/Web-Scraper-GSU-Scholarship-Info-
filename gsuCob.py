import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font

URL = "https://www.gram.edu/finaid/scholarships/privatescholar.php"
r = requests.get(URL)

soup = BeautifulSoup(r.content, 'html5lib')


myHeaders = []
headings = soup.select(".interior-main-description p")
for heading in headings:
    header = heading.text
    if ')' in header:
        bracketIndex = header.index(')')
        mainHeader = header[bracketIndex+1:]

        if 'http' in mainHeader:
            http = mainHeader.index('http')
            beforeHttp = mainHeader[:http].strip()
            myHeaders.append(beforeHttp)

        elif ':' in mainHeader:
            colon = mainHeader.index(':')
            beforeColon = mainHeader[:colon].strip()
            myHeaders.append(beforeColon)

        else:
            myHeaders.append(mainHeader.strip())

myParagraphs = []
paragraphs = soup.select(".interior-main-description p a")

for paragraph in paragraphs:
    myParagraphs.append(paragraph['href'])

myParagraphs.pop(58)


scholarships = 'Scholarships.xlsx'


'''
for i in range(len(myHeaders)):
    data = [myHeaders[i], myParagraphs[i]]
'''

wb = load_workbook(scholarships)
sheet = wb['Sheet1']

scholarshipsHead = sheet.cell(1, 1)
scholarshipsHead.value = 'Scholarships'
scholarshipsHead.font = Font(b=True)
linksHead = sheet.cell(1, 2)
linksHead.value = 'Link to apply'
linksHead.font = Font(b=True)

for row, head in enumerate(myHeaders):
    cell = sheet.cell(row+2, 1)
    cell.value = head

for row, link in enumerate(myParagraphs):
    cell = sheet.cell(row+2, 2)
    cell.value = link

# just to keep it nicely formatted
# (the links are very long, so Im filling the next line with white space)
for row, space in enumerate(myParagraphs):
    cell = sheet.cell(row+2, 3)
    cell.value = ' '

sheet.column_dimensions['A'].width = 75
sheet.column_dimensions['B'].width = 75

wb.save('my_Scholarships.xlsx')
